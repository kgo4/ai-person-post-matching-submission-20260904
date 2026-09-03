from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
rows=[
['AI平台Java后端工程师','北京/上海','3-5年','Java、Spring Boot、MySQL、Redis、Kafka、Python、RAG、Milvus、Docker','负责AI平台微服务、知识库检索、模型服务网关与高并发接口建设。','本科','熟悉大模型API、向量检索、异步任务与容器化部署'],
['智能系统Java开发工程师','上海/杭州','3-5年','Java、Spring Boot、MyBatis、PostgreSQL、Python、PyTorch、Elasticsearch、RabbitMQ','负责智能客服或推荐系统的业务服务、模型推理接口、缓存和监控建设。','本科','具备机器学习应用落地和模型服务工程化经验'],
['高级Java大模型应用架构师','深圳/北京','5-8年','Java、Spring Cloud、DDD、Netty、Agent工作流、RAG、Neo4j、Kubernetes','负责企业级Agent平台架构、多智能体编排、知识图谱和结构化输出校验。','本科及以上','具备高并发系统设计、平台治理和团队技术带领经验'],
['Java AI应用开发工程师','杭州/广州','2-4年','Java、Spring Boot、Python、OCR、RAG、Embedding、Redis、RabbitMQ','负责智能问答、文档解析、检索增强和异步批处理功能开发。','本科','熟悉Prompt工程、接口测试和AI能力业务集成'],
['智能制造Java平台工程师','广州/深圳','4-6年','Java、Netty、MQTT、WebSocket、Kafka、Flink、Python、InfluxDB','负责物联网设备数据平台、实时告警、预测性维护和模型推理服务。','本科','熟悉实时数据处理、时序数据和生产系统性能优化'],
]
wb=Workbook(); ws=wb.active; ws.title='岗位JD'
headers=['岗位名称','工作地点','经验要求','核心技能','岗位职责','学历要求','加分项']
ws.append(headers)
for r in rows: ws.append(r)
for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(horizontal='center')
for row in ws.iter_rows():
 for cell in row: cell.alignment=Alignment(wrap_text=True,vertical='top')
widths=[28,16,12,55,72,16,45]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
for i in range(2,ws.max_row+1): ws.row_dimensions[i].height=58
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
wb.save(r'C:\Users\pengchao\Downloads\Java_AI_岗位JD.xlsx')
